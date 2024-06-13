from openpyxl import load_workbook

compiled_file = "HC_compilado.xlsx"
raw_file = "HC Brasil 2023.06.xlsx"

destination_workbook = load_workbook(compiled_file)
destination_sheet = destination_workbook.active

raw_workbook = load_workbook(raw_file)
raw_sheet = raw_workbook.active

def get_headers():
    # TODO Tratar planilhas com 1 ou 2 linhas de cabeçalho
    return [cell.value for cell in destination_sheet]

def read_raw_HC():
    raw_headers = [cell.value for cell in raw_sheet[1]]
    correct_headers = [cell.value for cell in destination_sheet[1]]

    exported_headers = {}
  
    for i, value in enumerate(raw_headers):
        if value in correct_headers:
            exported_headers[i] = value
            print(f'{i}:{value} encontrado')
        else:
            print(f'{i}: Nāo encontrado')
 
    return(exported_headers)



def write_compiled_HC(header_map):
    # start_row = destination_sheet.max_row + 1
    
    for row in raw_sheet.iter_rows(min_row=2, values_only=True):
        new_row = [None] * len(destination_sheet[1]) #lista com quantidade de novas linhas
        
        for raw_key, raw_header in header_map.items():
            if raw_header in corrected_headers:
                final_index = corrected_headers.index(raw_header)
                new_row[final_index] = row[raw_key]
        
        destination_sheet.append(new_row)

header_map = read_raw_HC()
corrected_headers = [cell.value for cell in destination_sheet[1]]
write_compiled_HC(header_map)

# teste em 3625 linhas

destination_workbook.save(compiled_file)